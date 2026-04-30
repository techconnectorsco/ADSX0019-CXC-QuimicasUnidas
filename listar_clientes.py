"""
listar_clientes.py
Lista TODOS los clientes con paginación y exporta a Excel.
"""

import sys
sys.path.insert(0, '.')
from modules.database.conexion import ServiceLayerConnection
from datetime import datetime

# Intentar importar openpyxl, si no está disponible usar CSV
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    TIENE_EXCEL = True
except ImportError:
    TIENE_EXCEL = False
    print("⚠️ openpyxl no instalado, se generará CSV. Instalar con: pip install openpyxl")


def obtener_todos_paginado(conn, entidad, params, campo_orden="CardCode"):
    """
    Obtiene TODOS los registros usando paginación.
    El Service Layer de SAP tiene límite de 20 registros por consulta.
    """
    todos = []
    skip = 0
    page_size = 20  # Límite del servidor
    
    # Asegurar que tenga orderby para paginación consistente
    params["$orderby"] = campo_orden
    params["$top"] = page_size
    
    while True:
        params["$skip"] = skip
        resultado = conn.get(entidad, params)
        
        if not resultado or 'value' not in resultado or len(resultado['value']) == 0:
            break
        
        todos.extend(resultado['value'])
        skip += page_size
        
        # Mostrar progreso cada 100 registros
        if len(todos) % 100 == 0:
            print(f"   ... {len(todos)} registros obtenidos")
    
    return todos


def main():
    conn = ServiceLayerConnection(use_test_db=False)

    if not conn.login():
        print("❌ Error de conexión")
        return

    print("="*80)
    print("📋 EXTRACCIÓN COMPLETA DE CLIENTES")
    print(f"   Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("="*80)

    # =========================================================================
    # 1. Obtener vendedores (para nombres)
    # =========================================================================
    print("\n📥 Obteniendo vendedores...")
    vendedores_raw = obtener_todos_paginado(conn, "SalesPersons", {
        "$select": "SalesEmployeeCode,SalesEmployeeName,Active"
    }, "SalesEmployeeCode")
    
    vendedores_dict = {}
    for v in vendedores_raw:
        vendedores_dict[v.get('SalesEmployeeCode')] = v.get('SalesEmployeeName', '')
    print(f"   ✅ {len(vendedores_dict)} vendedores")

    # =========================================================================
    # 2. Obtener TODOS los clientes
    # =========================================================================
    print("\n📥 Obteniendo clientes (esto puede tomar unos segundos)...")
    
    clientes = obtener_todos_paginado(conn, "BusinessPartners", {
        "$filter": "CardType eq 'cCustomer'",
        "$select": "CardCode,CardName,Valid,SalesPersonCode,U_ZGIRA,CurrentAccountBalance,EmailAddress,Phone1,U_NVT_CorreoEstadoCuenta,U_NTV_EnvioAutomatico,CreditLimit"
    })
    
    print(f"   ✅ {len(clientes)} clientes obtenidos")

    conn.logout()

    # =========================================================================
    # 3. Procesar datos
    # =========================================================================
    print("\n🔄 Procesando datos...")
    
    # Separar activos e inactivos
    activos = [c for c in clientes if c.get('Valid') == 'tYES']
    inactivos = [c for c in clientes if c.get('Valid') != 'tYES']
    con_saldo = [c for c in activos if (c.get('CurrentAccountBalance') or 0) > 0]
    
    print(f"   Total: {len(clientes)}")
    print(f"   Activos: {len(activos)}")
    print(f"   Inactivos: {len(inactivos)}")
    print(f"   Con saldo > 0: {len(con_saldo)}")

    # =========================================================================
    # 4. Generar Excel o CSV
    # =========================================================================
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    
    if TIENE_EXCEL:
        archivo = f"clientes_quimicas_unidas_{fecha_str}.xlsx"
        generar_excel(clientes, vendedores_dict, archivo)
    else:
        archivo = f"clientes_quimicas_unidas_{fecha_str}.csv"
        generar_csv(clientes, vendedores_dict, archivo)
    
    print(f"\n✅ Archivo generado: {archivo}")
    print("="*80)


def limpiar_texto(valor):
    """Limpia caracteres ilegales para Excel."""
    if valor is None:
        return ""
    if not isinstance(valor, str):
        return valor
    # Eliminar caracteres de control (0x00-0x1F excepto tab, newline, carriage return)
    import re
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', valor)


def generar_excel(clientes, vendedores_dict, archivo):
    """Genera archivo Excel con formato."""
    print(f"\n📊 Generando Excel: {archivo}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0B114B", end_color="0B114B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        "Código", "Nombre", "Activo", "Vendedor Cód", "Vendedor Nombre", 
        "Zona Gira", "Saldo", "Email Principal", "Email CXC", 
        "Envío Auto", "Límite Crédito", "Teléfono"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos
    for row, c in enumerate(clientes, 2):
        vend_code = c.get('SalesPersonCode')
        vend_nombre = vendedores_dict.get(vend_code, '') if vend_code and vend_code != -1 else ''
        
        datos = [
            limpiar_texto(c.get('CardCode', '')),
            limpiar_texto(c.get('CardName', '')),
            'Sí' if c.get('Valid') == 'tYES' else 'No',
            vend_code if vend_code and vend_code != -1 else '',
            limpiar_texto(vend_nombre),
            limpiar_texto(str(c.get('U_ZGIRA', '') or '')),
            c.get('CurrentAccountBalance', 0) or 0,
            limpiar_texto(c.get('EmailAddress', '') or ''),
            limpiar_texto(c.get('U_NVT_CorreoEstadoCuenta', '') or ''),
            limpiar_texto(c.get('U_NTV_EnvioAutomatico', '') or ''),
            c.get('CreditLimit', 0) or 0,
            limpiar_texto(c.get('Phone1', '') or '')
        ]
        
        for col, valor in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.border = thin_border
            
            # Formato especial para saldo y límite crédito
            if col in [7, 11]:  # Saldo y Límite
                cell.number_format = '#,##0.00'
    
    # Ajustar anchos de columna
    anchos = [12, 45, 8, 12, 25, 10, 15, 35, 35, 10, 15, 15]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[chr(64 + col)].width = ancho
    
    # Filtros
    ws.auto_filter.ref = ws.dimensions
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    wb.save(archivo)
    print(f"   ✅ Excel guardado con {len(clientes)} registros")


def generar_csv(clientes, vendedores_dict, archivo):
    """Genera archivo CSV como alternativa."""
    import csv
    
    print(f"\n📊 Generando CSV: {archivo}")
    
    headers = [
        "Código", "Nombre", "Activo", "Vendedor Cód", "Vendedor Nombre", 
        "Zona Gira", "Saldo", "Email Principal", "Email CXC", 
        "Envío Auto", "Límite Crédito", "Teléfono"
    ]
    
    with open(archivo, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        
        for c in clientes:
            vend_code = c.get('SalesPersonCode')
            vend_nombre = vendedores_dict.get(vend_code, '') if vend_code and vend_code != -1 else ''
            
            writer.writerow([
                c.get('CardCode', ''),
                c.get('CardName', ''),
                'Sí' if c.get('Valid') == 'tYES' else 'No',
                vend_code if vend_code and vend_code != -1 else '',
                vend_nombre,
                c.get('U_ZGIRA', ''),
                c.get('CurrentAccountBalance', 0) or 0,
                c.get('EmailAddress', ''),
                c.get('U_NVT_CorreoEstadoCuenta', ''),
                c.get('U_NTV_EnvioAutomatico', ''),
                c.get('CreditLimit', 0) or 0,
                c.get('Phone1', '')
            ])
    
    print(f"   ✅ CSV guardado con {len(clientes)} registros")


if __name__ == "__main__":
    main()