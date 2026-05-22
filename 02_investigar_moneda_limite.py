"""
==========================================================================
SCRIPT 2 — INVESTIGACIÓN: MONEDA DEL LÍMITE DE CRÉDITO
==========================================================================
Investiga puntos B1, B2, B3 de los comentarios de Tania:

B1. C0326 Soluciones Tropicales: límite es CRC 15M pero gira lo muestra como USD
B2. C0470 Mobile Tech: límite es USD ($28,000) pero CXC lo muestra como CRC
B3. General: "Parece que no está tomando bien la información de qué cliente
    está en $ o colones"

OBJETIVO:
- Identificar QUÉ campo de SAP indica la moneda del límite de crédito
- Comparar contra lo que estamos asumiendo en agentes.py (que hardcodea CRC
  como default)
- Generar listado de clientes problemáticos

USO:
    python 02_investigar_moneda_limite.py
"""

import sys
from typing import List, Dict
from datetime import datetime

from modules.database.conexion import ServiceLayerConnection

# Clientes específicos reportados por Tania
CLIENTES_REPORTADOS = {
    "C0326": "Soluciones Tropicales — límite CRC 15M, gira lo mostró USD",
    "C0470": "Mobile Tech — límite USD $28,000, CXC lo mostró CRC",
}


def obtener_cliente_completo(conn, card_code: str) -> Dict:
    """Trae TODOS los campos de un cliente para inspeccionar qué hay disponible."""
    resp = conn.session.get(
        f"{conn.base_url}/BusinessPartners('{card_code}')",
        verify=False,
    )
    if not resp.ok:
        print(f"  ERROR al consultar {card_code}: {resp.status_code}")
        return None
    return resp.json()


def listar_campos_moneda(cliente: Dict):
    """Imprime todos los campos del cliente que tengan que ver con moneda o crédito."""
    relevantes = []
    for key, value in cliente.items():
        key_lower = key.lower()
        if any(
            palabra in key_lower
            for palabra in [
                "currenc",
                "moneda",
                "credit",
                "limit",
                "balance",
                "fc",
                "sys",
            ]
        ):
            relevantes.append((key, value))

    print(f"\n  Campos relevantes para moneda/crédito:")
    for k, v in sorted(relevantes):
        print(f"    {k:<40} = {v}")


def analizar_clientes_problematicos(conn):
    """Inspecciona los clientes específicos reportados por Tania."""
    print("=" * 70)
    print("INSPECCIÓN DETALLADA DE CLIENTES REPORTADOS")
    print("=" * 70)

    for code, descripcion in CLIENTES_REPORTADOS.items():
        print(f"\n>>> {code}: {descripcion}")
        cliente = obtener_cliente_completo(conn, code)
        if not cliente:
            continue

        print(f"\n  Nombre: {cliente.get('CardName', '')}")
        print(f"  CardCode: {cliente.get('CardCode', '')}")
        print(f"  CardType: {cliente.get('CardType', '')}")
        listar_campos_moneda(cliente)


def buscar_campo_moneda_universal(conn):
    """
    Pide UN cliente cualquiera con todos sus campos y lista TODOS los campos
    que puedan tener que ver con moneda. Esto sirve para descubrir el campo correcto.
    """
    print("\n" + "=" * 70)
    print("DESCUBRIMIENTO: Qué campos de moneda existen en BusinessPartners")
    print("=" * 70)

    resp = conn.session.get(
        f"{conn.base_url}/BusinessPartners",
        params={"$top": 1, "$filter": "CardType eq 'cCustomer'"},
        verify=False,
    )
    if not resp.ok:
        print(f"ERROR: {resp.status_code}")
        return

    data = resp.json()
    if not data.get("value"):
        print("No se obtuvieron clientes")
        return

    cliente = data["value"][0]
    print(
        f"\nCliente de muestra: {cliente.get('CardCode')} - {cliente.get('CardName')}"
    )
    print(f"\nTODOS los campos que mencionan moneda/crédito/balance:")
    print("-" * 70)
    for key, value in sorted(cliente.items()):
        key_lower = key.lower()
        if any(
            palabra in key_lower
            for palabra in ["currenc", "moneda", "credit", "limit", "balance", "fc"]
        ):
            print(f"  {key:<45} = {value}")


def listar_todos_con_currency(conn):
    """
    Lista TODOS los clientes con saldo distinto de 0 mostrando:
    - Currency (campo principal)
    - CreditLimit (en moneda local CRC)
    - CreditLimitFC (si existe, en moneda extranjera USD)
    - CurrentAccountBalance (CRC)
    - CurrentAccountBalanceFC (USD)
    """
    print("\n" + "=" * 70)
    print("LISTADO COMPLETO: Moneda y saldos de TODOS los clientes con saldo != 0")
    print("=" * 70)

    select = ",".join(
        [
            "CardCode",
            "CardName",
            "Currency",
            "CreditLimit",
            "MaxCommitment",
            "CurrentAccountBalance",
            "CurrentAccountBalanceFC",
            "CurrentAccountBalanceSys",
            "U_NTV_EnvioAutomatico",
        ]
    )
    filter_q = "CardType eq 'cCustomer' and CurrentAccountBalance ne 0"

    todos = []
    skip = 0
    page_size = 100
    while True:
        params = {
            "$select": select,
            "$filter": filter_q,
            "$top": page_size,
            "$skip": skip,
        }
        resp = conn.session.get(
            f"{conn.base_url}/BusinessPartners",
            params=params,
            verify=False,
        )
        if not resp.ok:
            print(f"ERROR: {resp.status_code}")
            break
        data = resp.json()
        batch = data.get("value", [])
        if not batch:
            break
        todos.extend(batch)
        if len(batch) < page_size:
            break
        skip += page_size

    print(f"\nTotal clientes consultados: {len(todos)}")

    # Distribución por moneda
    print("\nDistribución por campo 'Currency':")
    distribucion = {}
    for c in todos:
        cur = c.get("Currency") or "(vacío)"
        distribucion[cur] = distribucion.get(cur, 0) + 1
    for cur, cant in sorted(distribucion.items(), key=lambda x: -x[1]):
        print(f"  {cur:<15} -> {cant} clientes")

    # Imprimir tabla resumida (primeros 30 + clientes problemáticos)
    print("\n" + "-" * 110)
    print(
        f"{'Code':<8} {'Currency':<10} {'CreditLimit':>15} {'Balance CRC':>15} "
        f"{'Balance USD':>15} {'Envío':<6} Nombre"
    )
    print("-" * 110)

    # Primero los reportados por Tania
    problematicos = [c for c in todos if c["CardCode"] in CLIENTES_REPORTADOS]
    for c in problematicos:
        print_fila(c)

    print("  ... resto (primeros 30) ...")
    for c in todos[:30]:
        if c["CardCode"] not in CLIENTES_REPORTADOS:
            print_fila(c)

    return todos


def print_fila(c):
    envio = c.get("U_NTV_EnvioAutomatico") or "-"
    print(
        f"{c['CardCode']:<8} "
        f"{(c.get('Currency') or '-'):<10} "
        f"{(c.get('CreditLimit') or 0):>15,.2f} "
        f"{(c.get('CurrentAccountBalance') or 0):>15,.2f} "
        f"{(c.get('CurrentAccountBalanceFC') or 0):>15,.2f} "
        f"{envio:<6} "
        f"{(c.get('CardName') or '')[:40]}"
    )


def generar_reporte(clientes):
    """Genera Excel con todos los clientes y sus monedas."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("Instalar openpyxl para generar Excel: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Moneda Límite"
    headers = [
        "CardCode",
        "CardName",
        "Currency",
        "CreditLimit",
        "Balance CRC",
        "Balance USD",
        "Envío",
        "Sospechoso?",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="475DA4")

    for c in clientes:
        currency = c.get("Currency") or ""
        balance_crc = c.get("CurrentAccountBalance") or 0
        balance_usd = c.get("CurrentAccountBalanceFC") or 0

        # Heurística de sospecha:
        # - Si Currency dice "##" (multimoneda) o está vacío, sospechoso
        # - Si tiene Balance USD pero Currency dice CRC, sospechoso
        sospecha = ""
        if currency in ("##", "", None):
            sospecha = "Currency multimoneda o vacío"
        elif balance_usd != 0 and currency == "CRC":
            sospecha = "Tiene Balance USD pero Currency = CRC"

        ws.append(
            [
                c.get("CardCode"),
                c.get("CardName"),
                currency,
                c.get("CreditLimit") or 0,
                balance_crc,
                balance_usd,
                c.get("U_NTV_EnvioAutomatico") or "",
                sospecha,
            ]
        )

    salida = f"reporte_02_moneda_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(salida)
    print(f"\n✓ Reporte guardado: {salida}")


def main():
    print("=" * 70)
    print("SCRIPT 2 - INVESTIGACIÓN DE MONEDA DEL LÍMITE")
    print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    conn = ServiceLayerConnection(use_test_db=False)
    if not conn.login():
        print("ERROR: No se pudo conectar a SAP")
        sys.exit(1)

    # 1. Inspeccionar clientes reportados
    analizar_clientes_problematicos(conn)

    # 2. Descubrir todos los campos de moneda
    buscar_campo_moneda_universal(conn)

    # 3. Listar todos con su moneda
    clientes = listar_todos_con_currency(conn)

    # 4. Generar reporte
    generar_reporte(clientes)

    print("\n" + "=" * 70)
    print("INVESTIGACIÓN COMPLETADA")
    print("=" * 70)
    print("Comparte conmigo:")
    print(
        "1. La salida de consola (especialmente los campos relevantes de C0326 y C0470)"
    )
    print("2. El archivo Excel generado")


if __name__ == "__main__":
    main()
