"""
==========================================================================
SCRIPT 4 — INVESTIGACIÓN: PLAZOS DE CRÉDITO INCORRECTOS
==========================================================================
Investiga punto D1 de los comentarios de Tania:

D1. C0179 (El Colono Zona Norte) y C0257 (Mercado Materiales Tanque) tienen
    45 días de plazo en sistema pero la gira muestra 30 días.

OBJETIVO:
- Identificar cómo SAP almacena el plazo de pago
- Comparar el plazo en SAP vs lo que agentes.py está mostrando
- Determinar si:
    a) Estamos leyendo mal el campo
    b) Estamos hardcodeando un default (30) cuando no hay valor
    c) Hay un campo nuevo que no estamos consultando

USO:
    python 04_investigar_plazos.py
"""

import sys
from typing import List, Dict
from datetime import datetime

from modules.database.conexion import ServiceLayerConnection

# Clientes reportados por Tania
CLIENTES_INVESTIGAR = {
    "C0179": "El Colono Zona Norte (Excel: 45 días, Gira: 30)",
    "C0257": "Mercado Materiales Tanque (Excel: 45 días)",
}

# Otros clientes con plazos conocidos del Excel para validar
CLIENTES_REFERENCIA = {
    "C0012": ("AGRO UJARRAS", "Crédito a 30 días"),
    "C0138": ("EL LAGAR", "Crédito a 60 días"),
    "C0104": ("COOPEGUANACASTE", "Crédito a 90 días"),
    "C0048": ("ASOC. CAMARA PROD CAÑA", "Crédito a 45 días"),
}


def obtener_grupos_pago(conn):
    """
    Lista todos los grupos de condiciones de pago configurados en SAP.
    En SAP B1, los plazos vienen como GroupNum y se mapean en la tabla OCTG.
    El endpoint Service Layer es PaymentTermsTypes.
    """
    print("=" * 70)
    print("GRUPOS DE CONDICIONES DE PAGO EN SAP")
    print("=" * 70)

    resp = conn.session.get(
        f"{conn.base_url}/PaymentTermsTypes",
        params={"$select": "GroupNumber,PaymentTermsGroupName,NumberOfAdditionalDays"},
        verify=False,
    )
    if not resp.ok:
        print(f"ERROR: {resp.status_code}")
        return {}

    grupos = {}
    for g in resp.json().get("value", []):
        grupos[g["GroupNumber"]] = g
        print(
            f"  GroupNumber {g['GroupNumber']:>3}: "
            f"{g.get('PaymentTermsGroupName', ''):<30} "
            f"+ {g.get('NumberOfAdditionalDays', 0)} días"
        )

    return grupos


def inspeccionar_cliente(conn, card_code: str, descripcion: str, grupos_pago: Dict):
    """Trae el cliente y muestra el plazo de pago según SAP."""
    print(f"\n>>> {card_code} — {descripcion}")

    resp = conn.session.get(
        f"{conn.base_url}/BusinessPartners('{card_code}')",
        params={"$select": "CardCode,CardName,PayTermsGrpCode,ExtraDaysForPayment"},
        verify=False,
    )
    if not resp.ok:
        print(f"  ERROR: {resp.status_code}")
        return

    bp = resp.json()
    grp_code = bp.get("PayTermsGrpCode")
    extra_dias = bp.get("ExtraDaysForPayment", 0)

    print(f"  CardName: {bp.get('CardName')}")
    print(f"  PayTermsGrpCode: {grp_code}")
    print(f"  ExtraDaysForPayment: {extra_dias}")

    if grp_code in grupos_pago:
        grupo = grupos_pago[grp_code]
        print(
            f"  >>> Plazo según grupo {grp_code}: "
            f"{grupo.get('PaymentTermsGroupName')} "
            f"({grupo.get('NumberOfAdditionalDays', 0)} días)"
        )
    else:
        print(f"  >>> Grupo {grp_code} NO encontrado en lista de grupos")


def replicar_logica_actual(conn, card_code: str):
    """
    Replica cómo agentes.py / main.py está extrayendo el plazo
    según el código que vimos.
    """
    print(f"\n  --- Replicando lógica de extracción actual para {card_code} ---")

    # La función obtener_condicion_pago en main.py probablemente:
    # 1. Lee PayTermsGrpCode del cliente
    # 2. Lo mapea a un texto

    resp = conn.session.get(
        f"{conn.base_url}/BusinessPartners('{card_code}')",
        params={"$select": "PayTermsGrpCode"},
        verify=False,
    )
    if not resp.ok:
        return

    grp_code = resp.json().get("PayTermsGrpCode")
    print(f"  Código de grupo leído: {grp_code}")
    print(f"  Si en main.py hay un mapeo manual tipo {{1: '30 días', 2: '45 días'}}")
    print(f"  asegúrate que el código {grp_code} esté correctamente mapeado.")


def listar_todos_los_plazos(conn, grupos_pago: Dict):
    """
    Lista todos los clientes con saldo, mostrando su plazo según SAP.
    Útil para Tania como validación masiva.
    """
    print("\n" + "=" * 70)
    print("LISTADO COMPLETO DE PLAZOS POR CLIENTE")
    print("=" * 70)

    filter_q = "CardType eq 'cCustomer' and CurrentAccountBalance ne 0"
    todos = []
    skip = 0
    page_size = 100
    while True:
        params = {
            "$select": "CardCode,CardName,PayTermsGrpCode,ExtraDaysForPayment",
            "$filter": filter_q,
            "$top": page_size,
            "$skip": skip,
        }
        resp = conn.session.get(
            f"{conn.base_url}/BusinessPartners",
            params=params,
            verify=False,
        )
        if not resp.ok:
            break
        data = resp.json()
        batch = data.get("value", [])
        if not batch:
            break
        todos.extend(batch)
        if len(batch) < page_size:
            break
        skip += page_size

    print(f"Total clientes: {len(todos)}\n")

    # Distribución por código de grupo
    distribucion = {}
    for c in todos:
        grp = c.get("PayTermsGrpCode")
        distribucion[grp] = distribucion.get(grp, 0) + 1

    print("Distribución por código de plazo:")
    for grp, cant in sorted(distribucion.items()):
        nombre = grupos_pago.get(grp, {}).get("PaymentTermsGroupName", "???")
        dias = grupos_pago.get(grp, {}).get("NumberOfAdditionalDays", "???")
        print(f"  Código {grp:<5}: {cant:>4} clientes  -> {nombre} ({dias} días)")

    return todos


def generar_reporte(clientes, grupos_pago):
    """Genera Excel con plazos para que Tania revise."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Plazos"
    ws.append(
        ["CardCode", "CardName", "PayTermsGrpCode", "Grupo SAP", "Días", "ExtraDays"]
    )
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="475DA4")

    for c in clientes:
        grp = c.get("PayTermsGrpCode")
        nombre_grupo = grupos_pago.get(grp, {}).get("PaymentTermsGroupName", "???")
        dias = grupos_pago.get(grp, {}).get("NumberOfAdditionalDays", "???")
        ws.append(
            [
                c.get("CardCode"),
                c.get("CardName"),
                grp,
                nombre_grupo,
                dias,
                c.get("ExtraDaysForPayment", 0),
            ]
        )

    salida = f"reporte_04_plazos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(salida)
    print(f"\n✓ Reporte guardado: {salida}")


def main():
    print("=" * 70)
    print("SCRIPT 4 - INVESTIGACIÓN DE PLAZOS DE CRÉDITO")
    print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    conn = ServiceLayerConnection(use_test_db=False)
    if not conn.login():
        print("ERROR: No se pudo conectar a SAP")
        sys.exit(1)

    # 1. Listar todos los grupos de pago
    grupos = obtener_grupos_pago(conn)

    # 2. Inspeccionar clientes reportados por Tania
    print("\n" + "=" * 70)
    print("CASOS REPORTADOS POR TANIA")
    print("=" * 70)
    for code, desc in CLIENTES_INVESTIGAR.items():
        inspeccionar_cliente(conn, code, desc, grupos)
        replicar_logica_actual(conn, code)

    # 3. Clientes de referencia conocidos
    print("\n" + "=" * 70)
    print("CLIENTES DE REFERENCIA (validación)")
    print("=" * 70)
    for code, (nombre, plazo_esperado) in CLIENTES_REFERENCIA.items():
        inspeccionar_cliente(
            conn, code, f"{nombre} (esperado: {plazo_esperado})", grupos
        )

    # 4. Listado masivo
    clientes = listar_todos_los_plazos(conn, grupos)

    # 5. Reporte Excel
    generar_reporte(clientes, grupos)


if __name__ == "__main__":
    main()
