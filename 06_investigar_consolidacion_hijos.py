"""
==========================================================================
SCRIPT 6 — INVESTIGACIÓN: CONSOLIDACIÓN DE CUENTAS HIJAS EN UN SOLO PDF
==========================================================================
Investiga puntos F2 y F3 de los comentarios de Tania:

F2. C0223 Jotocillo tiene varias cuentas hijas (C0223, C0224, C0225) que
    deberían salir en UN solo PDF al cliente, no 3 correos separados.
    Tania adjuntó el PDF de muestra (Antigüedad_de_saldos_de_clientes...)
    donde se ve que SAP B1 nativo SÍ consolida.
F3. Lo mismo aplica para Colono Agropecuario y Almacenes Colono.

OBJETIVO:
- Descubrir QUÉ campo en SAP indica la relación padre-hijo entre cuentas
- Confirmar la relación para los grupos reportados
- Generar lista de TODOS los grupos padre-hijo encontrados, para que en
  el refactor el sistema agrupe correctamente

CAMPOS CANDIDATOS EN SAP B1:
- FatherCard: el CardCode del padre (campo estándar de SAP)
- ConsolidatingBP: ¿se consolida con otro?
- ProjectCode: a veces se usa para agrupar
- Algún UDF custom

USO:
    python 06_investigar_consolidacion_hijos.py
"""

import sys
from typing import List, Dict
from datetime import datetime
from collections import defaultdict

from modules.database.conexion import ServiceLayerConnection

# Grupos reportados por Tania que DEBERÍAN consolidarse
GRUPOS_TANIA = [
    {
        "padre_esperado": "C0223",
        "miembros": ["C0223", "C0224", "C0225"],
        "nombre": "INVERSIONES JOTOCILLO",
    },
    {
        "padre_esperado": "C0138",
        "miembros": ["C0138", "C0139", "C0140"],
        "nombre": "EL LAGAR",
    },
    {
        "padre_esperado": "C0161",
        "miembros": ["C0161", "C0162", "C0163", "C0164"],
        "nombre": "COLONO AGROPECUARIO",
    },
    {
        "padre_esperado": "C0040",
        "miembros": ["C0040", "C0042", "C0043"],
        "nombre": "ALMACENES COLONO",
    },
    {
        "padre_esperado": "C0346",
        "miembros": ["C0346", "C0347"],
        "nombre": "UNICOMER GOLLO",
    },
    {
        "padre_esperado": "C0314",
        "miembros": ["C0314", "C0315", "C0316"],
        "nombre": "CARLOS RUIZ",
    },
    {
        "padre_esperado": "C0398",
        "miembros": ["C0398", "C0499"],
        "nombre": "FRANCISCO JIMENEZ (según notas del Excel)",
    },
]


def buscar_campo_relacion(conn, card_code: str):
    """
    Trae TODOS los campos de un cliente y busca cualquiera que sugiera
    una relación con otro cliente (padre, consolidación, agrupación).
    """
    print(f"\n>>> Inspeccionando todos los campos de {card_code}...")

    resp = conn.session.get(
        f"{conn.base_url}/BusinessPartners('{card_code}')",
        verify=False,
    )
    if not resp.ok:
        print(f"  ERROR: {resp.status_code}")
        return

    bp = resp.json()
    print(f"  Nombre: {bp.get('CardName')}")
    print(f"\n  Campos que sugieren relación padre-hijo o agrupación:")
    print("-" * 70)

    palabras_clave = [
        "father",
        "parent",
        "consolidat",
        "main",
        "group",
        "padre",
        "hijo",
        "head",
        "master",
        "linked",
        "associated",
        "related",
        "proyecto",
        "project",
    ]

    for key, value in sorted(bp.items()):
        key_lower = key.lower()
        if any(p in key_lower for p in palabras_clave):
            # Solo mostrar si tiene valor o si es un campo prometedor
            if value not in (None, "", 0, "tNO"):
                print(f"  ✓ {key:<40} = {value}")
            else:
                print(f"    {key:<40} = {value}")


def verificar_grupo(conn, grupo: Dict):
    """Verifica un grupo padre-hijo de los reportados por Tania."""
    print(f"\n{'='*70}")
    print(f"GRUPO: {grupo['nombre']}")
    print(f"  Miembros esperados: {', '.join(grupo['miembros'])}")
    print(f"  Padre esperado:     {grupo['padre_esperado']}")
    print(f"{'='*70}")

    for code in grupo["miembros"]:
        resp = conn.session.get(
            f"{conn.base_url}/BusinessPartners('{code}')",
            params={
                "$select": "CardCode,CardName,FatherCard,ConsolBP,GroupCode,U_Padre,U_NTV_Padre"
            },
            verify=False,
        )
        if not resp.ok:
            # Intentar sin los campos custom que pueden no existir
            resp = conn.session.get(
                f"{conn.base_url}/BusinessPartners('{code}')",
                params={"$select": "CardCode,CardName,FatherCard,ConsolBP,GroupCode"},
                verify=False,
            )

        if not resp.ok:
            print(f"  {code}: ERROR {resp.status_code}")
            continue

        bp = resp.json()
        print(f"\n  {code} — {bp.get('CardName', '')[:50]}")
        print(f"    FatherCard:  {bp.get('FatherCard') or '(vacío)'}")
        print(f"    ConsolBP:    {bp.get('ConsolBP') or '(vacío)'}")
        print(f"    GroupCode:   {bp.get('GroupCode') or '(vacío)'}")
        print(f"    U_Padre:     {bp.get('U_Padre') or '(no existe el campo)'}")
        print(f"    U_NTV_Padre: {bp.get('U_NTV_Padre') or '(no existe el campo)'}")


def buscar_padres_globalmente(conn):
    """
    Trae TODOS los clientes y agrupa los que tengan el mismo FatherCard
    (o el campo equivalente que descubramos).
    """
    print("\n" + "=" * 70)
    print("BÚSQUEDA GLOBAL DE GRUPOS PADRE-HIJO")
    print("=" * 70)

    resp = conn.session.get(
        f"{conn.base_url}/BusinessPartners",
        params={
            "$select": "CardCode,CardName,FatherCard,ConsolBP,GroupCode",
            "$filter": "CardType eq 'cCustomer'",
            "$top": 1000,
        },
        verify=False,
    )

    if not resp.ok:
        print(f"ERROR: {resp.status_code}")
        return

    clientes = resp.json().get("value", [])
    print(f"Total clientes analizados: {len(clientes)}")

    # Agrupar por FatherCard
    por_father = defaultdict(list)
    por_consol = defaultdict(list)
    sin_padre = []

    for c in clientes:
        father = c.get("FatherCard")
        consol = c.get("ConsolBP")
        if father:
            por_father[father].append(c["CardCode"])
        elif consol:
            por_consol[consol].append(c["CardCode"])
        else:
            sin_padre.append(c["CardCode"])

    print(f"\nClientes con FatherCard:  {sum(len(v) for v in por_father.values())}")
    print(f"Clientes con ConsolBP:    {sum(len(v) for v in por_consol.values())}")
    print(f"Clientes sin agrupación:  {len(sin_padre)}")

    if por_father:
        print(f"\n--- Grupos detectados por FatherCard ---")
        for padre, hijos in sorted(por_father.items()):
            nombre_padre = next(
                (c["CardName"] for c in clientes if c["CardCode"] == padre), "???"
            )
            print(f"  Padre {padre} ({nombre_padre[:30]}): {len(hijos)} hijos")
            print(
                f"    Hijos: {', '.join(hijos[:10])}{'...' if len(hijos) > 10 else ''}"
            )

    if por_consol:
        print(f"\n--- Grupos detectados por ConsolBP ---")
        for cabeza, miembros in sorted(por_consol.items()):
            nombre = next(
                (c["CardName"] for c in clientes if c["CardCode"] == cabeza), "???"
            )
            print(f"  Cabeza {cabeza} ({nombre[:30]}): {len(miembros)} miembros")
            print(
                f"    Miembros: {', '.join(miembros[:10])}{'...' if len(miembros) > 10 else ''}"
            )


def generar_reporte(conn):
    """Genera reporte Excel de grupos padre-hijo."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return

    resp = conn.session.get(
        f"{conn.base_url}/BusinessPartners",
        params={
            "$select": "CardCode,CardName,FatherCard,ConsolBP,GroupCode,CurrentAccountBalance",
            "$filter": "CardType eq 'cCustomer' and CurrentAccountBalance ne 0",
            "$top": 500,
        },
        verify=False,
    )
    if not resp.ok:
        return

    clientes = resp.json().get("value", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Grupos Padre-Hijo"
    ws.append(
        ["CardCode", "CardName", "FatherCard", "ConsolBP", "GroupCode", "Balance"]
    )
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="475DA4")

    for c in sorted(
        clientes, key=lambda x: (x.get("FatherCard") or x["CardCode"], x["CardCode"])
    ):
        ws.append(
            [
                c.get("CardCode"),
                c.get("CardName"),
                c.get("FatherCard", ""),
                c.get("ConsolBP", ""),
                c.get("GroupCode", ""),
                c.get("CurrentAccountBalance", 0),
            ]
        )

    salida = f"reporte_06_padre_hijo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(salida)
    print(f"\n✓ Reporte guardado: {salida}")


def main():
    print("=" * 70)
    print("SCRIPT 6 - INVESTIGACIÓN DE CONSOLIDACIÓN PADRE-HIJO")
    print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    conn = ServiceLayerConnection(use_test_db=False)
    if not conn.login():
        print("ERROR: No se pudo conectar a SAP")
        sys.exit(1)

    # 1. Inspeccionar todos los campos del primer cliente reportado
    # para descubrir qué campo está usando SAP para la relación
    print("\n[Paso 1] Buscar el campo de relación padre-hijo")
    buscar_campo_relacion(conn, "C0224")  # Hijo esperado de C0223
    buscar_campo_relacion(conn, "C0139")  # Hijo esperado de C0138

    # 2. Verificar cada grupo reportado
    print("\n[Paso 2] Verificar grupos reportados por Tania")
    for grupo in GRUPOS_TANIA:
        verificar_grupo(conn, grupo)

    # 3. Búsqueda global
    print("\n[Paso 3] Búsqueda global de relaciones padre-hijo")
    buscar_padres_globalmente(conn)

    # 4. Reporte
    generar_reporte(conn)

    print("\n" + "=" * 70)
    print("INVESTIGACIÓN COMPLETADA")
    print("=" * 70)
    print("Comparte conmigo la salida de consola completa.")
    print("Especialmente para identificar:")
    print("  - Qué campo de SAP indica la relación padre-hijo")
    print("  - Si los grupos esperados por Tania efectivamente están")
    print("    relacionados en SAP, o si la relación está solo en otro lado")


if __name__ == "__main__":
    main()
