"""
==========================================================================
SCRIPT 1 — INVESTIGACIÓN: FILTROS Y AUDIENCIA DE CXC
==========================================================================
Investiga puntos A1, A2, A3, A4, A5 y F1 de los comentarios de Tania:

A1. Clientes marcados NO en sistema pero recibieron CXC
    (Coopeguanacaste, Evonik, Colono Agrop, Unicomer)
A2. Empleados/gerencias que recibieron CXC sin deber
    (Mario Vargas, Allan Rodríguez, Roland, Jens, "Químicas", Axel)
A3. Unicomer Gollo C0346-347 no salió en gira de ningún agente
A4. Clientes con NO pero sin correo registrado que "toman correo de otro lado"
A5. Conteo: Tania espera 93, procesamos 75. Diferencia.
F1. C0138-39-40 y C0223-24-25 no salieron aunque tienen SÍ y correos

USO:
    Colocar este archivo en la RAÍZ del proyecto Quimicas-Unidas/
    Asegurarse de tener el Excel de Tania en la misma carpeta:
        CLIENTES_ENVIO_DE_ESTADOS_DE_CUENTA_20-05-2026.xlsx

    Luego correr:
        python 01_investigar_filtros_y_audiencia.py

SALIDA:
    - Imprime resumen en consola
    - Genera reporte_01_audiencia.xlsx con los detalles
"""

import sys
import os
from typing import List, Dict, Set
from datetime import datetime

# Tu conexión existente
from modules.database.conexion import ServiceLayerConnection

# Para leer el Excel de Tania
try:
    import openpyxl
except ImportError:
    print("ERROR: necesitas instalar openpyxl -> pip install openpyxl")
    sys.exit(1)


EXCEL_TANIA = "CLIENTES_ENVIO_DE_ESTADOS_DE_CUENTA_20-05-2026.xlsx"


# =============================================================================
# PASO 1: Leer el Excel de Tania
# =============================================================================


def leer_excel_tania(ruta: str) -> Dict[str, Dict]:
    """
    Lee el Excel y devuelve un dict {CardCode: {datos del Excel}}
    Procesa AMBAS secciones del Excel (activos y "CON SALDOS EN CUENTA/INACTIVOS")
    """
    if not os.path.exists(ruta):
        print(f"ERROR: No se encuentra el archivo {ruta}")
        print(f"Cópialo a la carpeta actual: {os.getcwd()}")
        sys.exit(1)

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.active

    clientes = {}
    seccion_actual = "activos"
    headers = None
    headers_inactivos = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Primera fila = encabezados
        if row_idx == 1:
            headers = list(row)
            continue

        # Detectar línea separadora de la sección de inactivos
        if row[0] == "CON SALDOS EN CUENTA/ INACTIVOS":
            seccion_actual = "inactivos"
            continue

        # En la sección de inactivos, la siguiente fila son los headers
        if seccion_actual == "inactivos" and row[0] == "Código SN":
            headers_inactivos = list(row)
            continue

        # Filas de totales al final
        if row[0] is None or (isinstance(row[0], str) and not row[0].startswith("C")):
            continue

        # Procesar fila de cliente
        h = (
            headers_inactivos
            if (seccion_actual == "inactivos" and headers_inactivos)
            else headers
        )
        card_code = row[0]
        if not card_code or not str(card_code).startswith("C"):
            continue

        cliente = {}
        for i, col_name in enumerate(h):
            if i < len(row):
                cliente[col_name] = row[i]
        cliente["_seccion"] = seccion_actual
        clientes[str(card_code)] = cliente

    wb.close()
    return clientes


# =============================================================================
# PASO 2: Leer maestro de clientes de SAP
# =============================================================================


def obtener_todos_clientes_sap(conn: ServiceLayerConnection) -> List[Dict]:
    """
    Trae TODOS los clientes (Business Partners tipo cliente con saldo > 0)
    desde SAP B1, sin filtros, para ver el universo completo.
    """
    print("\n[SAP] Consultando maestro completo de clientes...")

    select = ",".join(
        [
            "CardCode",
            "CardName",
            "CardType",
            "EmailAddress",
            "Phone1",
            "Phone2",
            "Cellular",
            "CurrentAccountBalance",
            "SalesPersonCode",
            "U_ZGIRA",
            "U_NVT_CorreoEstadoCuenta",
            "U_NTV_EnvioAutomatico",
            "CreditLimit",
            "Currency",
            "PayTermsGrpCode",
            "FreeText",
            "Frozen",
            "Valid",
        ]
    )

    # Solo clientes (CardType = 'C' o 'cCustomer'), con saldo distinto de 0
    filter_q = "CardType eq 'cCustomer' and CurrentAccountBalance ne 0"

    todos = []
    skip = 0
    page_size = 100

    while True:
        params = {
            "$select": select,
            "$filter": filter_q,
            "$top": page_size,
            "$skip": skip,
        }
        resp = conn.session.get(
            f"{conn.base_url}/BusinessPartners",
            params=params,
            verify=False,
        )
        if not resp.ok:
            print(f"ERROR en consulta SAP: {resp.status_code} - {resp.text[:300]}")
            break

        data = resp.json()
        batch = data.get("value", [])
        if not batch:
            break
        todos.extend(batch)
        if len(batch) < page_size:
            break
        skip += page_size
        print(f"  ... {len(todos)} clientes leídos")

    print(f"[SAP] Total: {len(todos)} clientes con saldo distinto de 0")
    return todos


# =============================================================================
# PASO 3: Analizar discrepancias
# =============================================================================


def normalizar_si_no(valor) -> str:
    """Normaliza valores 'Si'/'No'/'Sí'/'SI'/None a 'SI' o 'NO' o 'VACIO'"""
    if valor is None or valor == "":
        return "VACIO"
    v = str(valor).strip().upper()
    if v in ("SI", "SÍ", "S", "Y", "YES"):
        return "SI"
    if v in ("NO", "N"):
        return "NO"
    return f"OTRO:{v}"


def analizar(clientes_excel: Dict[str, Dict], clientes_sap: List[Dict]):
    """Compara Excel de Tania vs SAP y reporta discrepancias."""

    # Indexar SAP por CardCode
    sap_dict = {c["CardCode"]: c for c in clientes_sap}

    # Conjuntos
    codes_excel = set(clientes_excel.keys())
    codes_sap = set(sap_dict.keys())

    print("\n" + "=" * 70)
    print("RESUMEN GENERAL")
    print("=" * 70)
    print(f"Clientes en Excel de Tania:       {len(codes_excel)}")
    print(f"Clientes en SAP con saldo != 0:   {len(codes_sap)}")
    print(f"En Excel pero NO en SAP:          {len(codes_excel - codes_sap)}")
    print(f"En SAP pero NO en Excel:          {len(codes_sap - codes_excel)}")
    print(f"En ambos:                          {len(codes_excel & codes_sap)}")

    # =========================================================================
    # CASO 1: Comparar el campo "Envío automático"
    # =========================================================================
    print("\n" + "=" * 70)
    print("COMPARACIÓN: Excel 'Envio automatico' vs SAP 'U_NTV_EnvioAutomatico'")
    print("=" * 70)

    discrepancias_envio = []
    for code in sorted(codes_excel & codes_sap):
        exc = normalizar_si_no(
            clientes_excel[code].get("Envio automatico de Estado de cuenta")
        )
        sap = normalizar_si_no(sap_dict[code].get("U_NTV_EnvioAutomatico"))
        if exc != sap:
            discrepancias_envio.append(
                {
                    "CardCode": code,
                    "CardName": sap_dict[code].get("CardName", ""),
                    "excel_envio": exc,
                    "sap_envio": sap,
                    "excel_correo_ec": clientes_excel[code].get(
                        "Correo Estado de Cuenta", ""
                    ),
                    "sap_correo_ec": sap_dict[code].get("U_NVT_CorreoEstadoCuenta", ""),
                    "sap_email_general": sap_dict[code].get("EmailAddress", ""),
                }
            )

    print(f"Total discrepancias: {len(discrepancias_envio)}")
    print("\nPrimeras 20 discrepancias:")
    print(f"{'Code':<8} {'Excel':<8} {'SAP':<10} Nombre")
    print("-" * 70)
    for d in discrepancias_envio[:20]:
        print(
            f"{d['CardCode']:<8} {d['excel_envio']:<8} {d['sap_envio']:<10} {d['CardName'][:50]}"
        )

    # =========================================================================
    # CASO 2: Clientes que Tania reportó como problemáticos
    # =========================================================================
    print("\n" + "=" * 70)
    print("CASOS PUNTUALES REPORTADOS POR TANIA")
    print("=" * 70)

    casos_tania = {
        "C0104": "Coopeguanacaste (debe estar NO)",
        "C0040": "Almacenes Colono (debe estar NO)",
        "C0161": "El Colono Agropecuario (debe estar NO)",
        "C0162": "Colono Agropecuario (hijo)",
        "C0163": "Colono Agropecuario (hijo)",
        "C0164": "Colono Agropecuario (hijo)",
        "C0042": "Almacenes Colono (hijo)",
        "C0043": "Almacenes Colono (hijo)",
        "C0346": "Unicomer Gollo (debe estar NO)",
        "C0347": "Unicomer Gollo (debe estar NO)",
        "C0138": "El Lagar (Tania dice SI y NO salió)",
        "C0139": "El Lagar (hijo)",
        "C0140": "El Lagar (hijo)",
        "C0223": "Jotocillo (debe consolidar con hijos)",
        "C0224": "Jotocillo (hijo)",
        "C0225": "Jotocillo (hijo)",
    }

    for code, descripcion in casos_tania.items():
        en_excel = code in codes_excel
        en_sap = code in codes_sap
        envio_excel = (
            normalizar_si_no(
                clientes_excel.get(code, {}).get("Envio automatico de Estado de cuenta")
            )
            if en_excel
            else "N/A"
        )
        envio_sap = (
            normalizar_si_no(sap_dict.get(code, {}).get("U_NTV_EnvioAutomatico"))
            if en_sap
            else "N/A"
        )
        correo_sap = (
            sap_dict.get(code, {}).get("U_NVT_CorreoEstadoCuenta", "") if en_sap else ""
        )
        email_general_sap = (
            sap_dict.get(code, {}).get("EmailAddress", "") if en_sap else ""
        )
        saldo = sap_dict.get(code, {}).get("CurrentAccountBalance", 0) if en_sap else 0

        print(f"\n{code} - {descripcion}")
        print(f"  En Excel: {en_excel}    En SAP: {en_sap}    Saldo SAP: {saldo:,.2f}")
        print(f"  Envío -> Excel: {envio_excel}    SAP: {envio_sap}")
        print(f"  Correo EC (SAP U_NVT_CorreoEstadoCuenta): {correo_sap or '(VACÍO)'}")
        print(
            f"  EmailAddress general SAP:                 {email_general_sap or '(VACÍO)'}"
        )

    # =========================================================================
    # CASO 3: Empleados / Gerencias / Cuenta "Químicas"
    # =========================================================================
    print("\n" + "=" * 70)
    print("CLIENTES CON NOMBRE DE EMPLEADO / GERENCIA")
    print("=" * 70)
    print("(Buscando en SAP por CardName que contenga apellidos o palabras clave)")

    palabras_clave = [
        "MARIO VARGAS",
        "ALLAN RODRIGUEZ",
        "ROLAND",
        "JENS",
        "QUIMICAS UNIDAS",
        "AXEL",
    ]
    posibles_empleados = []
    for c in clientes_sap:
        nombre_upper = (c.get("CardName") or "").upper()
        for kw in palabras_clave:
            if kw in nombre_upper:
                posibles_empleados.append(
                    {
                        "CardCode": c["CardCode"],
                        "CardName": c.get("CardName"),
                        "envio": normalizar_si_no(c.get("U_NTV_EnvioAutomatico")),
                        "correo_ec": c.get("U_NVT_CorreoEstadoCuenta", ""),
                        "email_general": c.get("EmailAddress", ""),
                        "saldo": c.get("CurrentAccountBalance", 0),
                    }
                )
                break

    print(f"Encontrados: {len(posibles_empleados)}")
    for e in posibles_empleados:
        print(
            f"  {e['CardCode']} | {e['CardName'][:40]:<40} | Envío: {e['envio']:<6} | Saldo: {e['saldo']:>14,.2f}"
        )
        print(
            f"    Correo EC: {e['correo_ec'] or '(vacío)'}    Email general: {e['email_general'] or '(vacío)'}"
        )

    # =========================================================================
    # CASO 4: Conteo final que esperaría procesar el sistema
    # =========================================================================
    print("\n" + "=" * 70)
    print("CONTEO ESPERADO SEGÚN CADA FUENTE (con saldo != 0)")
    print("=" * 70)

    # Según Excel
    excel_si = sum(
        1
        for c in clientes_excel.values()
        if normalizar_si_no(c.get("Envio automatico de Estado de cuenta")) == "SI"
    )
    excel_no = sum(
        1
        for c in clientes_excel.values()
        if normalizar_si_no(c.get("Envio automatico de Estado de cuenta")) == "NO"
    )

    # Según SAP
    sap_si = sum(
        1
        for c in clientes_sap
        if normalizar_si_no(c.get("U_NTV_EnvioAutomatico")) == "SI"
    )
    sap_no = sum(
        1
        for c in clientes_sap
        if normalizar_si_no(c.get("U_NTV_EnvioAutomatico")) == "NO"
    )
    sap_vacio = sum(
        1
        for c in clientes_sap
        if normalizar_si_no(c.get("U_NTV_EnvioAutomatico")) == "VACIO"
    )

    print(f"Excel — SI:    {excel_si}")
    print(f"Excel — NO:    {excel_no}")
    print(f"Excel — TOTAL: {len(clientes_excel)}")
    print()
    print(f"SAP   — SI:    {sap_si}")
    print(f"SAP   — NO:    {sap_no}")
    print(f"SAP   — VACÍO: {sap_vacio}")
    print(f"SAP   — TOTAL: {len(clientes_sap)}")
    print()
    print(f"Tania reporta que el sistema procesó: 75")
    print(f"Tania espera para envío automático:    85")

    return discrepancias_envio, posibles_empleados


# =============================================================================
# PASO 4: Generar reporte Excel
# =============================================================================


def generar_reporte_excel(
    discrepancias_envio, posibles_empleados, clientes_excel, clientes_sap
):
    """Crea reporte_01_audiencia.xlsx con todos los detalles para Tania."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # ---- Hoja 1: Discrepancias de Envío SI/NO ----
    ws1 = wb.active
    ws1.title = "Discrepancias Envío"
    headers = [
        "CardCode",
        "CardName",
        "Excel Envío",
        "SAP Envío",
        "Correo EC (Excel)",
        "Correo EC (SAP)",
        "Email General (SAP)",
    ]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="475DA4")

    for d in discrepancias_envio:
        ws1.append(
            [
                d["CardCode"],
                d["CardName"],
                d["excel_envio"],
                d["sap_envio"],
                d["excel_correo_ec"],
                d["sap_correo_ec"],
                d["sap_email_general"],
            ]
        )

    # ---- Hoja 2: Posibles empleados/gerencias ----
    ws2 = wb.create_sheet("Empleados-Gerencias")
    ws2.append(["CardCode", "CardName", "Envío", "Saldo", "Correo EC", "Email General"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="DC3545")

    for e in posibles_empleados:
        ws2.append(
            [
                e["CardCode"],
                e["CardName"],
                e["envio"],
                e["saldo"],
                e["correo_ec"],
                e["email_general"],
            ]
        )

    # ---- Hoja 3: Clientes solo en SAP (no en Excel) ----
    ws3 = wb.create_sheet("Solo en SAP")
    ws3.append(
        ["CardCode", "CardName", "Envío SAP", "Saldo", "Correo EC", "Email General"]
    )
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="288FCC")

    codes_excel = set(clientes_excel.keys())
    for c in clientes_sap:
        if c["CardCode"] not in codes_excel:
            ws3.append(
                [
                    c["CardCode"],
                    c.get("CardName", ""),
                    normalizar_si_no(c.get("U_NTV_EnvioAutomatico")),
                    c.get("CurrentAccountBalance", 0),
                    c.get("U_NVT_CorreoEstadoCuenta", ""),
                    c.get("EmailAddress", ""),
                ]
            )

    # ---- Hoja 4: Clientes solo en Excel (no en SAP) ----
    ws4 = wb.create_sheet("Solo en Excel")
    ws4.append(["CardCode", "CardName", "Envío Excel", "Correo EC"])
    for cell in ws4[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="28A745")

    codes_sap = set(c["CardCode"] for c in clientes_sap)
    for code, c in clientes_excel.items():
        if code not in codes_sap:
            ws4.append(
                [
                    code,
                    c.get("Nombre SN", ""),
                    normalizar_si_no(c.get("Envio automatico de Estado de cuenta")),
                    c.get("Correo Estado de Cuenta", ""),
                ]
            )

    salida = f"reporte_01_audiencia_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(salida)
    print(f"\n✓ Reporte guardado: {salida}")
    return salida


# =============================================================================
# MAIN
# =============================================================================


def main():
    print("=" * 70)
    print("SCRIPT 1 - INVESTIGACIÓN DE FILTROS Y AUDIENCIA")
    print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    # 1. Leer Excel
    print(f"\n[1/3] Leyendo Excel: {EXCEL_TANIA}")
    clientes_excel = leer_excel_tania(EXCEL_TANIA)
    print(f"  Cargados {len(clientes_excel)} clientes del Excel")

    # 2. Conectar a SAP y leer maestro
    print(f"\n[2/3] Conectando a SAP B1...")
    conn = ServiceLayerConnection(use_test_db=False)
    if not conn.login():
        print("ERROR: No se pudo conectar a SAP")
        sys.exit(1)
    print("  Conectado.")

    clientes_sap = obtener_todos_clientes_sap(conn)

    # 3. Analizar
    print(f"\n[3/3] Analizando discrepancias...")
    discrepancias_envio, posibles_empleados = analizar(clientes_excel, clientes_sap)

    # 4. Generar reporte
    generar_reporte_excel(
        discrepancias_envio, posibles_empleados, clientes_excel, clientes_sap
    )

    print("\n" + "=" * 70)
    print("INVESTIGACIÓN COMPLETADA")
    print("=" * 70)
    print("Revisa el archivo Excel generado y compártelo conmigo")
    print("para definir los siguientes pasos.")


if __name__ == "__main__":
    main()
