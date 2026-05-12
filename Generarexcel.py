"""
generarexcel.py - Químicas Unidas
Generación de Excel para Estados de Cuenta.

Características:
- Mismo formato de datos que generarpdf.py
- Tabla unificada (Primero USD, luego CRC)
- Formato de moneda Latinoamericano (puntos para miles, comas para decimales)
- Estilos profesionales con colores acordes al PDF
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict
import os

# =============================================================================
# CONSTANTES Y CONFIGURACIÓN DE COLORES
# =============================================================================

AZUL_OSCURO = "0B114B"  # Header, títulos
AZUL_CLARO = "288FCC"  # Línea decorativa
AZUL_FOOTER = "475DA4"  # Fondo footer
ROJO = "DC3545"  # Vencidos
VERDE = "28A745"  # Al día
GRIS = "646464"  # Textos secundarios
GRIS_CLARO = "F5F5F5"  # Fondo alternado

# =============================================================================
# FUNCIONES AUXILIARES
# =============================================================================


def formato_latino(valor: float) -> str:
    """
    Convierte el formato US (1,234.56) a formato Latino (1.234,56).
    Es 100% seguro sin importar el idioma del sistema operativo.
    """
    if valor is None:
        valor = 0.0
    num_str = f"{abs(valor):,.2f}"
    num_str = num_str.replace(",", "X")
    num_str = num_str.replace(".", ",")
    num_str = num_str.replace("X", ".")
    return num_str


# =============================================================================
# CLASE PRINCIPAL - EXCEL ESTADO DE CUENTA
# =============================================================================


class ExcelEstadoCuenta:
    """Genera Excel de Estado de Cuenta con estilos profesionales."""

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Estado de Cuenta"
        self.row = 1

        # Configurar ancho de página para impresión
        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_LEGAL
        self.ws.page_setup.orientation = "landscape"
        self.ws.print_options.horizontalCentered = True

        # Márgenes para impresión
        self.ws.page_margins.left = 0.5
        self.ws.page_margins.right = 0.5
        self.ws.page_margins.top = 0.75
        self.ws.page_margins.bottom = 0.75

    def _obtener_fill(self, hex_color: str) -> PatternFill:
        """Crea un PatternFill a partir de un color hex."""
        return PatternFill(
            start_color=hex_color, end_color=hex_color, fill_type="solid"
        )

    def _obtener_border(self) -> Border:
        """Define el borde estándar para celdas."""
        thin_border = Side(style="thin", color="000000")
        return Border(
            left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
        )

    def _establecer_ancho_columnas(self):
        """Establece los anchos de columna según el contenido."""
        anchos = {
            "A": 18,  # No de Doc
            "B": 18,  # No de Orden
            "C": 18,  # Fecha Factura
            "D": 18,  # Fecha Vencimiento
            "E": 12,  # Trans
            "F": 45,  # Descripción
            "G": 22,  # Monto Factura
            "H": 16,  # Estatus
        }
        for col, ancho in anchos.items():
            self.ws.column_dimensions[col].width = ancho

    def agregar_encabezado(self):
        """Agrega el encabezado con logo, título, fecha y hora."""
        # Título principal
        self.ws.merge_cells("A1:H1")
        cell = self.ws["A1"]
        cell.value = "Químicas Unidas Ltda."
        cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        cell.fill = self._obtener_fill(AZUL_FOOTER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        self.ws.row_dimensions[1].height = 26

        # Subtítulo
        self.ws.merge_cells("A2:H2")
        cell = self.ws["A2"]
        cell.value = "Estado de Cuenta"
        cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        cell.fill = self._obtener_fill(AZUL_FOOTER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        self.ws.row_dimensions[2].height = 22

        # Fecha y hora
        self.ws.merge_cells("A3:H3")
        cell = self.ws["A3"]
        fecha = datetime.now().strftime("%d/%m/%Y")
        hora = datetime.now().strftime("%H:%M")
        cell.value = f"Fecha: {fecha}  |  Hora: {hora}"
        cell.font = Font(name="Arial", size=10, color="FFFFFF")
        cell.fill = self._obtener_fill(AZUL_CLARO)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        self.ws.row_dimensions[3].height = 18

        self.row = 5  # Espacio después del encabezado

    def agregar_datos_cliente(self, cliente: Dict):
        """Agrega la sección de datos del cliente."""
        border = self._obtener_border()
        fill_header = self._obtener_fill(AZUL_CLARO)

        # Fila 1: Cliente y Contacto
        self.ws.merge_cells(f"A{self.row}:B{self.row}")
        cell = self.ws[f"A{self.row}"]
        cell.value = "Cliente:"
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[self.row].height = 20

        self.ws.merge_cells(f"C{self.row}:D{self.row}")
        cell = self.ws[f"C{self.row}"]
        nombre_cliente = f"{cliente.get('codigo', '')} - {cliente.get('nombre', '')}"
        cell.value = nombre_cliente
        cell.font = Font(name="Arial", size=10, bold=False)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        self.ws.merge_cells(f"E{self.row}:F{self.row}")
        cell = self.ws[f"E{self.row}"]
        cell.value = "Contacto:"
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        self.ws.merge_cells(f"G{self.row}:H{self.row}")
        cell = self.ws[f"G{self.row}"]
        cell.value = cliente.get("contacto", "") or "No especificado"
        cell.font = Font(name="Arial", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        self.row += 1

        # Fila 2: Correo y Teléfono
        self.ws.merge_cells(f"A{self.row}:B{self.row}")
        cell = self.ws[f"A{self.row}"]
        cell.value = "Correo:"
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[self.row].height = 20

        self.ws.merge_cells(f"C{self.row}:D{self.row}")
        cell = self.ws[f"C{self.row}"]
        cell.value = cliente.get("correo", "") or "No registrado"
        cell.font = Font(name="Arial", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        self.ws.merge_cells(f"E{self.row}:F{self.row}")
        cell = self.ws[f"E{self.row}"]
        cell.value = "Teléfono:"
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        self.ws.merge_cells(f"G{self.row}:H{self.row}")
        cell = self.ws[f"G{self.row}"]
        cell.value = cliente.get("telefono", "") or "No registrado"
        cell.font = Font(name="Arial", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        self.row += 1

        # Fila 3: Dirección y Vendedor
        self.ws.merge_cells(f"A{self.row}:B{self.row}")
        cell = self.ws[f"A{self.row}"]
        cell.value = "Dirección:"
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[self.row].height = 20

        self.ws.merge_cells(f"C{self.row}:D{self.row}")
        cell = self.ws[f"C{self.row}"]
        cell.value = cliente.get("direccion", "") or "No registrada"
        cell.font = Font(name="Arial", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        self.ws.merge_cells(f"E{self.row}:F{self.row}")
        cell = self.ws[f"E{self.row}"]
        cell.value = "Vendedor:"
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        self.ws.merge_cells(f"G{self.row}:H{self.row}")
        cell = self.ws[f"G{self.row}"]
        cell.value = cliente.get("vendedor", "")
        cell.font = Font(name="Arial", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        self.row += 1

        # Fila 4: Condición de Pago y Límite de Crédito
        self.ws.merge_cells(f"A{self.row}:B{self.row}")
        cell = self.ws[f"A{self.row}"]
        cell.value = "Cond. Pago:"
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.row_dimensions[self.row].height = 20

        self.ws.merge_cells(f"C{self.row}:D{self.row}")
        cell = self.ws[f"C{self.row}"]
        cell.value = cliente.get("condicion_pago", "")
        cell.font = Font(name="Arial", size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        self.ws.merge_cells(f"E{self.row}:F{self.row}")
        cell = self.ws[f"E{self.row}"]
        cell.value = "Límite Crédito:"
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        self.ws.merge_cells(f"G{self.row}:H{self.row}")
        cell = self.ws[f"G{self.row}"]
        limite = cliente.get("limite_credito", 0)
        cell.value = f"CRC {formato_latino(limite)}"
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="right", vertical="center")

        self.row += 2  # Espacio en blanco

    def agregar_tabla_documentos(
        self, docs_usd: List[Dict], docs_crc: List[Dict], totales: Dict
    ):
        """Agrega la tabla de documentos (USD + CRC)."""
        todos_docs = []
        if docs_usd:
            todos_docs.extend(docs_usd)
        if docs_crc:
            todos_docs.extend(docs_crc)

        if not todos_docs:
            return

        headers = [
            "No de Doc",
            "No de Orden",
            "Fecha Factura",
            "Fecha Vencimiento",
            "Trans",
            "Descripción",
            "Monto Factura",
            "Estatus",
        ]

        # Encabezados
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.row, column=col_num)
            cell.value = header
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = self._obtener_fill(AZUL_FOOTER)
            cell.border = self._obtener_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")

        self.ws.row_dimensions[self.row].height = 18
        self.row += 1

        # Datos
        fila_par = False
        border = self._obtener_border()

        for doc in todos_docs:
            esta_vencido = doc.get("esta_vencido", False)
            saldo = doc.get("saldo", 0)
            moneda = doc.get("moneda", "")
            simbolo = "USD" if moneda == "USD" else "CRC"

            # Colores alternados
            fill_fondo = (
                self._obtener_fill("FAFAFA")
                if fila_par
                else self._obtener_fill("FFFFFF")
            )
            fila_par = not fila_par

            # No de Doc
            cell = self.ws.cell(row=self.row, column=1)
            consecutivo = doc.get("consecutivo_fe", "") or str(doc.get("doc_num", ""))
            cell.value = consecutivo
            cell.fill = fill_fondo
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=9)

            # No de Orden
            cell = self.ws.cell(row=self.row, column=2)
            orden = str(doc.get("orden_compra", ""))
            if len(orden) > 15:
                orden = orden[:12] + "..."
            cell.value = orden
            cell.fill = fill_fondo
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=9)

            # Fecha Factura
            cell = self.ws.cell(row=self.row, column=3)
            fecha = doc.get("fecha", "")
            if len(fecha) >= 10:
                try:
                    fecha = datetime.strptime(fecha[:10], "%Y-%m-%d").strftime(
                        "%d/%m/%Y"
                    )
                except:
                    pass
            cell.value = fecha
            cell.fill = fill_fondo
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=9)

            # Fecha Vencimiento
            cell = self.ws.cell(row=self.row, column=4)
            fecha_vence = doc.get("fecha_vence", "")
            if len(fecha_vence) >= 10:
                try:
                    fecha_vence = datetime.strptime(
                        fecha_vence[:10], "%Y-%m-%d"
                    ).strftime("%d/%m/%Y")
                except:
                    pass
            cell.value = fecha_vence
            cell.fill = fill_fondo
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=9)

            # Trans
            cell = self.ws.cell(row=self.row, column=5)
            trans = doc.get("tipo_codigo", "")
            cell.value = trans
            cell.fill = fill_fondo
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=9)

            # Descripción
            cell = self.ws.cell(row=self.row, column=6)
            desc = doc.get("descripcion", "")[:80]
            cell.value = desc
            cell.fill = fill_fondo
            cell.border = border
            cell.alignment = Alignment(horizontal="left", wrap_text=True)
            cell.font = Font(name="Arial", size=9)

            # Monto Factura
            cell = self.ws.cell(row=self.row, column=7)
            monto_str = f"{simbolo} {formato_latino(abs(saldo))}"
            if saldo < 0:
                monto_str = f"({monto_str})"
            cell.value = monto_str
            cell.fill = fill_fondo
            cell.border = border
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(name="Arial", size=9)

            # Estatus
            cell = self.ws.cell(row=self.row, column=8)
            if saldo < 0:
                estatus = "A favor"
                color_estatus = VERDE
            elif esta_vencido:
                estatus = "Vencido"
                color_estatus = ROJO
            else:
                estatus = "Al día"
                color_estatus = VERDE

            cell.value = estatus
            cell.fill = fill_fondo
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=9, bold=True, color=color_estatus)

            self.row += 1

        # Fila de totales
        self.ws.merge_cells(f"A{self.row}:G{self.row}")
        cell = self.ws[f"A{self.row}"]

        if docs_usd:
            cell.value = "TOTAL GENERAL USD:"
            cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            cell.fill = self._obtener_fill(AZUL_FOOTER)
            cell.border = self._obtener_border()
            cell.alignment = Alignment(horizontal="right", vertical="center")

            cell_total = self.ws.cell(row=self.row, column=8)
            cell_total.value = f"USD {formato_latino(totales['dolares'])}"
            cell_total.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            cell_total.fill = self._obtener_fill(AZUL_FOOTER)
            cell_total.border = self._obtener_border()
            cell_total.alignment = Alignment(horizontal="right")

            self.ws.row_dimensions[self.row].height = 20
            self.row += 1

        if docs_crc:
            self.ws.merge_cells(f"A{self.row}:G{self.row}")
            cell = self.ws[f"A{self.row}"]
            cell.value = "TOTAL GENERAL COLONES:"
            cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            cell.fill = self._obtener_fill(AZUL_FOOTER)
            cell.border = self._obtener_border()
            cell.alignment = Alignment(horizontal="right", vertical="center")

            cell_total = self.ws.cell(row=self.row, column=8)
            cell_total.value = f"CRC {formato_latino(totales['colones'])}"
            cell_total.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            cell_total.fill = self._obtener_fill(AZUL_FOOTER)
            cell_total.border = self._obtener_border()
            cell_total.alignment = Alignment(horizontal="right")

            self.ws.row_dimensions[self.row].height = 20
            self.row += 1

        self.row += 1  # Espacio en blanco

    def agregar_seccion_vencidos(self, rangos_usd: Dict, rangos_crc: Dict):
        """Agrega la sección de análisis de vencimiento."""
        vencido_usd = rangos_usd.get("total_vencido", 0) if rangos_usd else 0
        vencido_crc = rangos_crc.get("total_vencido", 0) if rangos_crc else 0

        if vencido_usd <= 0 and vencido_crc <= 0:
            return

        self.row += 1  # Espacio

        # Procesar USD
        if vencido_usd > 0:
            self._agregar_tabla_vencidos_moneda(
                rangos_usd, "USD", "Dólares", vencido_usd, col_inicio=1
            )

        # Procesar CRC
        if vencido_crc > 0:
            self._agregar_tabla_vencidos_moneda(
                rangos_crc, "CRC", "Colones", vencido_crc, col_inicio=5
            )

    def _agregar_tabla_vencidos_moneda(
        self, rangos: Dict, prefijo: str, titulo: str, total_v: float, col_inicio: int
    ):
        """Agrega una tabla de vencidos para una moneda específica."""
        border = self._obtener_border()

        # Título
        cell = self.ws.cell(row=self.row, column=col_inicio)
        self.ws.merge_cells(
            f"{get_column_letter(col_inicio)}{self.row}:{get_column_letter(col_inicio + 3)}{self.row}"
        )
        cell.value = f"Total Facturas Vencidas en {titulo}"
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = self._obtener_fill(AZUL_FOOTER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        self.ws.row_dimensions[self.row].height = 18

        self.row += 1

        # Encabezados de tabla
        labels = [
            ("Total 0-30", "0_30"),
            ("Total 31-60", "31_60"),
            ("Total 61-90", "61_90"),
            ("Total 91-120", "91_120"),
            ("Total 120+", "mas_120"),
        ]

        for label, key in labels:
            col1 = get_column_letter(col_inicio)
            col2 = get_column_letter(col_inicio + 3)
            self.ws.merge_cells(f"{col1}{self.row}:{col2}{self.row}")
            cell = self.ws.cell(row=self.row, column=col_inicio)
            cell.value = label
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = self._obtener_fill(AZUL_CLARO)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

            valor = rangos.get(key, 0)
            col_valor = col_inicio + 4
            cell_valor = self.ws.cell(row=self.row, column=col_valor)
            val_str = f"{prefijo} {formato_latino(valor)}" if valor > 0 else ""
            cell_valor.value = val_str
            cell_valor.font = Font(name="Arial", size=10)
            cell_valor.fill = self._obtener_fill("FFFFFF")
            cell_valor.border = border
            cell_valor.alignment = Alignment(horizontal="right")

            self.ws.row_dimensions[self.row].height = 18
            self.row += 1

        # Total vencido
        col1 = get_column_letter(col_inicio)
        col2 = get_column_letter(col_inicio + 3)
        self.ws.merge_cells(f"{col1}{self.row}:{col2}{self.row}")
        cell = self.ws.cell(row=self.row, column=col_inicio)
        cell.value = "TOTAL VENCIDO"
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = self._obtener_fill(AZUL_FOOTER)
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        col_valor = col_inicio + 4
        cell_valor = self.ws.cell(row=self.row, column=col_valor)
        cell_valor.value = f"{prefijo} {formato_latino(total_v)}"
        cell_valor.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell_valor.fill = self._obtener_fill(AZUL_FOOTER)
        cell_valor.border = border
        cell_valor.alignment = Alignment(horizontal="right")

        self.ws.row_dimensions[self.row].height = 18
        self.row += 2  # Espacio entre tablas

    def guardar(self, filepath: str) -> str:
        """Guarda el workbook en la ruta especificada."""
        self._establecer_ancho_columnas()
        self.wb.save(filepath)
        return filepath


# =============================================================================
# PUNTO DE ENTRADA
# =============================================================================


def generar_excel_estado_cuenta(
    datos: Dict, output_dir: str = "data/estados_cuenta"
) -> str:
    """
    Genera el Excel de estado de cuenta para un cliente.

    Args:
        datos: Dict con estructura de preparar_datos_cliente() del main.py
        output_dir: Directorio donde guardar el Excel

    Returns:
        Ruta del archivo Excel generado
    """
    os.makedirs(output_dir, exist_ok=True)

    # Crear Excel
    excel = ExcelEstadoCuenta()

    # Agregar componentes
    excel.agregar_encabezado()
    excel.agregar_datos_cliente(datos["cliente"])
    excel.agregar_tabla_documentos(
        datos["documentos"]["dolares"], datos["documentos"]["colones"], datos["totales"]
    )
    excel.agregar_seccion_vencidos(
        datos["rangos_vencimiento"].get("USD"), datos["rangos_vencimiento"].get("CRC")
    )

    # Generar nombre de archivo (mismo patrón que PDF)
    codigo = datos["cliente"]["codigo"]
    nombre = datos["cliente"]["nombre"].replace(" ", "_")[:20]
    fecha = datetime.now().strftime("%Y%m%d")
    filename = f"EC_{codigo}_{nombre}_{fecha}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # Guardar
    excel.guardar(filepath)

    return filepath
